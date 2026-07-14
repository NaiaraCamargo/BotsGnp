import os
import shutil
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Alignment

from pncp_shared.logs.controle_logs import logs
from pncp_shared.config.controle_config import config


def normalizar_texto(valor):
    if valor is None or pd.isna(valor):
        return ""

    return str(valor).strip()


def normalizar_numero(valor):
    try:
        if valor is None or pd.isna(valor):
            return 0

        if isinstance(valor, (int, float)):
            return float(valor)

        valor = str(valor).strip()
        valor = valor.replace("R$", "")
        valor = valor.replace(" ", "")
        valor = valor.replace(".", "")
        valor = valor.replace(",", ".")

        return float(valor)
    except Exception:
        return 0


def identificar_coluna_reserva(df, coluna_configurada=""):
    if coluna_configurada:
        for coluna in df.columns:
            if str(coluna).strip().lower() == coluna_configurada.strip().lower():
                return coluna

        raise Exception(f"\nColuna configurada não encontrada no Excel: {coluna_configurada}")

    nomes_possiveis = [
        "Reserva",
        "RESERVA"    
    ]

    for coluna in df.columns:
        nome_coluna = str(coluna).strip().lower()

        if nome_coluna in nomes_possiveis:
            return coluna

    raise Exception("\nNão foi possível identificar a coluna do numero da reserva no Excel.")


def carregar_premios_por_reserva(caminho_excel):
    try:
        df_banco = pd.read_excel(
            caminho_excel,
            sheet_name="CÓPIA BANCO TODOS LOTES MAPFRE",
            usecols="B:D",
            keep_default_na=False
        )

        if df_banco.empty:
            return {}

        coluna_reserva = df_banco.columns[0]
        coluna_premio = df_banco.columns[2]

        df_banco[coluna_reserva] = df_banco[coluna_reserva].apply(normalizar_texto)
        df_banco[coluna_premio] = df_banco[coluna_premio].apply(normalizar_numero)

        premios_por_reserva = (
            df_banco
            .groupby(coluna_reserva)[coluna_premio]
            .sum()
            .to_dict()
        )

        return premios_por_reserva

    except Exception as ex:
        logs.error(f"\nErro ao carregar premios por reserva: {ex}")
        return {}


def ler_links_excel(caminho_excel):
    try:
        caminho_excel = normalizar_texto(caminho_excel)
        aba_excel = config("aba_excel")

        if not caminho_excel:
            raise Exception("\nCaminho do Excel não informado no config.json.")

        if not os.path.exists(caminho_excel):
            raise Exception(f"\nArquivo Excel não encontrado: {caminho_excel}")

        df = pd.read_excel(caminho_excel, sheet_name=aba_excel, keep_default_na=False)

        if df.empty:
            logs.info("\nExcel está vazio.")
            return []

        reservas = []
        premios_por_reserva = carregar_premios_por_reserva(caminho_excel)

        for _, linha in df.iterrows():
            processados = normalizar_texto(linha.get("PROCESSADO", "")).upper()

            if processados == "OK":
                continue

            reserva = linha.get("RESERVA", "")
            ramo = linha.get("RAMO", "")
            premio = linha.get("PRÊMIO", "")
            reserva_normalizada = normalizar_texto(reserva)

            if normalizar_texto(premio) == "":
                premio = premios_por_reserva.get(reserva_normalizada, 0)
            
            reservas.append({
                "reserva": reserva_normalizada,
                "ramo": normalizar_texto(ramo),
                "premio": 0 if pd.isna(premio) else premio
            })


        logs.info(f"\nTotal de reservas válidas encontradas: {len(reservas)}")
        print(f"Total de reservas válidas encontradas: {len(reservas)}")

        return reservas

    except Exception as e:
        logs.error(f"\nErro ao ler links do Excel: {e}")
        raise


def salvar_workbook_com_retry(wb, caminho_excel, tentativas=5, espera_segundos=2):
    ultimo_erro = None

    for tentativa in range(1, tentativas + 1):
        try:
            wb.save(caminho_excel)
            return
        except PermissionError as ex:
            ultimo_erro = ex
            logs.error("\nERRO ao salvar planilha (tentativa %s/%s): %s", tentativa, tentativas, ex)
            if tentativa < tentativas:
                time.sleep(espera_segundos)
        except Exception:
            raise

    if ultimo_erro:
        raise ultimo_erro


def criar_backup_excel(caminho_excel, max_backups=3):
    try:
        caminho_excel = normalizar_texto(caminho_excel)
        caminho_original = os.path.abspath(caminho_excel)

        if not os.path.exists(caminho_original):
            return ""

        pasta_excel = os.path.dirname(caminho_original)
        nome_excel = os.path.splitext(os.path.basename(caminho_original))[0]
        pasta_backup = os.path.join(pasta_excel, "_backup_bot_mapfre")
        os.makedirs(pasta_backup, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        caminho_backup = os.path.join(pasta_backup, f"{nome_excel}_{timestamp}.xlsx")
        shutil.copy2(caminho_original, caminho_backup)

        backups = sorted(
            [os.path.join(pasta_backup, arquivo) for arquivo in os.listdir(pasta_backup) if arquivo.lower().endswith(".xlsx")],
            key=os.path.getmtime,
            reverse=True,
        )

        for backup_antigo in backups[max_backups:]:
            try:
                os.remove(backup_antigo)
            except Exception:
                pass

        logs.info(f"\nBackup da planilha criado: {caminho_backup}")
        return caminho_backup
    except Exception as ex:
        logs.error(f"\nERRO ao criar backup da planilha: {ex}")
        return ""

def atualizar_planilha(quantidade_arquivos, quantidade_baixado, orgao, link, caminho_excel, id_reserva, processou, msg):
    wb = None
    try:
        print("\nIniciado processo de atualização da planilha")
        caminho_excel = normalizar_texto(caminho_excel)
        aba_excel = config("aba_excel")

        if not caminho_excel:
            raise Exception("\nnaminho do Excel não informado no config.json.")

        if not os.path.exists(caminho_excel):
            raise Exception(f"\nArquivo Excel não encontrado: {caminho_excel}")

        criar_backup_excel(caminho_excel)
        wb = load_workbook(caminho_excel)
        ws = wb[aba_excel]
        
        col_id_reserva = obter_ou_criar_coluna(ws, "RESERVA")
        col_quantidade_baixados = obter_ou_criar_coluna(ws, "Nº ARQUIVOS BAIXADOS")
        col_quantidade_arquivos = obter_ou_criar_coluna(ws, "Nº ARQUIVOS TOTAL")
        col_data = obter_ou_criar_coluna(ws,"DATA ARQUIVOS BAIXADOS")
        col_link = obter_ou_criar_coluna(ws,"LINK RESERVA")
        col_orgao = obter_ou_criar_coluna(ws,"ÓRGÃO")
        col_processado = obter_ou_criar_coluna(ws, "PROCESSADO")
        col_msg = obter_ou_criar_coluna(ws, "MENSAGEM DO BOT" )
        
        if processou:
            processado = "OK"
        else:
            processado = "ERRO"
            
        linhas_encontradas = None
        
        id_reserva_normalizado = normalizar_texto(id_reserva).replace(".0", "")

        for row in range(2, ws.max_row + 1):
            valor_reserva = ws.cell(row=row, column=col_id_reserva).value
            
            if normalizar_texto(valor_reserva).replace(".0", "") == id_reserva_normalizado:
                linhas_encontradas = row
                print(f"Atualizando reserva numero {id_reserva_normalizado} na linha {linhas_encontradas} msg bot {msg}")
                logs.info(f"\nAtualizando reserva numero {id_reserva_normalizado} na linha {linhas_encontradas}")
                break
            
        if linhas_encontradas is None:
            raise ValueError(f"\nReserva {id_reserva} não encontrada no Excel.")
         
        gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_quantidade_baixados, quantidade_baixado)
        gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_quantidade_arquivos, quantidade_arquivos)
        gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_data, datetime.now().strftime("%d/%m/%Y %H:%M"))
        gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_link, link)
        gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_orgao, orgao, sobrescrever=False)
        gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_processado, processado)
        
        if processou:
            gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_msg, "")
        else:
            gravar_valor_com_estilo_primeira_coluna(ws, linhas_encontradas, col_msg, msg)
  
        salvar_workbook_com_retry(wb, caminho_excel)
        
        print("Finalizado processo de atualizaçao de planilha")
         
    except Exception as ex:
        logs.error(f"\nERRO ao atualizar planilha: {ex}")
        raise
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

def obter_ou_criar_coluna(ws, nome_coluna):
    try:
        nome_coluna = str(nome_coluna).strip()
        headers = {}
        
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1,column=col).value
            if valor:
                headers[str(valor).strip()] = col
                
        if nome_coluna in headers:
            return headers[nome_coluna]
        
        nova_coluna = ws.max_column + 1
        
        celula_base = ws.cell(row=1, column=2)
        celula_nova = ws.cell(row=1, column=nova_coluna)

        celula_nova.value = nome_coluna
        celula_nova.font = copy(celula_base.font)
        celula_nova.fill = copy(celula_base.fill)
        celula_nova.border = copy(celula_base.border)
        celula_nova.alignment = copy(celula_base.alignment)
        celula_nova.number_format = celula_base.number_format
        celula_nova.protection = copy(celula_base.protection)

        letra_nova = ws.cell(row=1, column=nova_coluna).column_letter
        letra_base = ws.cell(row=1, column=1).column_letter
        ws.column_dimensions[letra_nova].width = ws.column_dimensions[letra_base].width

        return nova_coluna
        
    except Exception as ex:
        logs.error(f"\nERRO ao obter ou criar a coluna: {ex}")
        raise

def gravar_valor_com_estilo_primeira_coluna(ws, row, col, valor, sobrescrever=True):
    try:
        if col <= 1:
            return False

        celula_destino = ws.cell(row=row, column=col)

        if not sobrescrever:
            valor_atual = celula_destino.value
            if valor_atual is not None and str(valor_atual).strip() != "":
                return False

        celula_base = ws.cell(row=2, column=2)

        celula_destino.value = valor
        celula_destino._style = copy(celula_base._style)
        celula_destino.alignment = Alignment(horizontal="center", vertical="center")
        
        return True

    except Exception as ex:
        logs.error(f"\nERRO ao gravar valor com estilo na planilha: {ex}")
        return False
